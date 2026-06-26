import os, json
from collections import defaultdict
from itertools import combinations
# reuse the Opta data + group constants from the exact matrix module
from opponent_matrix_exact import POSITIONS, QUAL, GROUPS, HOST_WINNERS
pos={t:[a/100,b/100,c/100,d/100] for g,t,a,b,c,d in POSITIONS}
teams_in=defaultdict(list)
for g,t,*_ in POSITIONS: teams_in[g].append(t)
q3={t:max(0.0,QUAL[t]-pos[t][0]-pos[t][1]) for t in pos}
A={g:sum(q3[t] for t in teams_in[g]) for g in GROUPS}
table={''.join(sorted(r['advancing'])):r['assign'] for r in json.load(open('assignment_table.json')).values()}
combos=[];Z=0.0
for eight in combinations(GROUPS,8):
    S=set(eight);w=1.0
    for g in GROUPS: w*= A[g] if g in S else (1-A[g])
    if w==0: continue
    combos.append((S,w,table[''.join(sorted(eight))]));Z+=w
combos=[(S,w/Z,amap) for S,w,amap in combos]
# hosted-third distribution when England (1L) win
hostL=defaultdict(float)
for S,w,amap in combos:
    Y=amap['1L'][1]
    for u in teams_in[Y]:
        if q3[u]>0: hostL[u]+=w*q3[u]/A[Y]
# England overall opponent (exact): P1 host + P2 -> 2K runner-up + P3 -> 1K winner
EP1,EP2,EP3=.8171,.1772,.0057
opp=defaultdict(float)
for u,p in hostL.items(): opp[u]+=EP1*p
for u in teams_in['K']: opp[u]+=EP2*pos[u][1]   # 2nd -> Group K runner-up
for u in teams_in['K']: opp[u]+=EP3*pos[u][0]   # 3rd -> Group K winner
opp=dict(sorted(opp.items(),key=lambda x:-x[1]))
hostL=dict(sorted(hostL.items(),key=lambda x:-x[1]))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
hdr=Font(bold=True,color='FFFFFF'); hf=PatternFill('solid',start_color='1D0A30')
wb=Workbook(); s=wb.active; s.title='England R32 Opponent'
s['A1']='England 2026 World Cup - Round-of-32 Opponent (EXACT, Opta-grounded)'; s['A1'].font=Font(bold=True,size=14)
s['A2']='No Monte Carlo. Opta finish-position marginals + FIFA 495-combination table (independent-Bernoulli joint).'
s.append([]); s.append(['Opponent','Probability'])
for c in s[4]: c.font=hdr;c.fill=hf
for u,p in opp.items():
    if p>=.0005: s.append([u,round(p,4)])
for r in range(5,5+sum(1 for p in opp.values() if p>=.0005)): s.cell(r,2).number_format='0.0%'
s.column_dimensions['A'].width=20;s.column_dimensions['B'].width=12

s2=wb.create_sheet('If England Win L')
s2['A1']='Third-placed team hosted in Match 80 (given England win Group L)';s2['A1'].font=Font(bold=True)
s2.append([]);s2.append(['Team','Probability'])
for c in s2[3]: c.font=hdr;c.fill=hf
for u,p in hostL.items():
    if p>=.0005: s2.append([u,round(p,4)])
for r in range(4,4+sum(1 for p in hostL.values() if p>=.0005)): s2.cell(r,2).number_format='0.0%'
s2.column_dimensions['A'].width=18;s2.column_dimensions['B'].width=12

s3=wb.create_sheet('Third Advancement A_g')
s3.append(['Group','P(3rd advances)']);
for c in s3[1]: c.font=hdr;c.fill=hf
for g in GROUPS: s3.append([g,round(A[g],4)])
s3.append(['SUM',round(sum(A.values()),4)])
for r in range(2,15): s3.cell(r,2).number_format='0.0%'
s3.column_dimensions['A'].width=10;s3.column_dimensions['B'].width=16

s4=wb.create_sheet('Finish Positions')
s4.append(['Group','Team','1st','2nd','3rd','4th','qual'])
for c in s4[1]: c.font=hdr;c.fill=hf
for g,t,a,b,c,d in POSITIONS: s4.append([g,t,a/100,b/100,c/100,d/100,QUAL[t]])
for r in range(2,2+len(POSITIONS)):
    for col in (3,4,5,6,7): s4.cell(r,col).number_format='0.0%'
s4.column_dimensions['A'].width=8;s4.column_dimensions['B'].width=16

out='England_R32_forecast_exact.xlsx'
try: wb.save(out)
except PermissionError:
    import time; out=f'England_R32_forecast_exact_{int(time.time())}.xlsx'; wb.save(out)
print('Saved',out)
print('England opponent (exact):')
for u,p in opp.items():
    if p>=.0005: print(f'  {u:14}{p*100:5.1f}%')
